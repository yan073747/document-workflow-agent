from __future__ import annotations

import csv
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.models import SalesAnalysis, SalesMetricItem


REQUIRED_COLUMNS = {
    "order_id",
    "date",
    "region",
    "salesperson",
    "product",
    "category",
    "quantity",
    "unit_price",
    "total_amount",
    "customer_type",
}


def load_sales_rows(file_path: str | Path) -> list[dict[str, Any]]:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File does not exist: {path}")
    if path.suffix.lower() == ".csv":
        rows = _load_csv(path)
    elif path.suffix.lower() == ".xlsx":
        rows = _load_xlsx(path)
    else:
        raise ValueError("Only CSV and XLSX files are supported in the first version.")

    if not rows:
        raise ValueError("Sales file is empty.")

    missing = REQUIRED_COLUMNS - set(rows[0].keys())
    if missing:
        missing_text = ", ".join(sorted(missing))
        raise ValueError(f"Sales file is missing required columns: {missing_text}")
    return rows


def analyze_sales_file(file_path: str | Path) -> SalesAnalysis:
    rows = load_sales_rows(file_path)
    order_ids: set[str] = set()
    total_revenue = 0.0
    total_quantity = 0
    revenue_by_region: dict[str, float] = defaultdict(float)
    quantity_by_region: dict[str, int] = defaultdict(int)
    revenue_by_category: dict[str, float] = defaultdict(float)
    quantity_by_category: dict[str, int] = defaultdict(int)
    revenue_by_salesperson: dict[str, float] = defaultdict(float)
    quantity_by_salesperson: dict[str, int] = defaultdict(int)
    revenue_by_month: dict[str, float] = defaultdict(float)
    quantity_by_month: dict[str, int] = defaultdict(int)
    revenue_by_customer: dict[str, float] = defaultdict(float)

    for row in rows:
        order_id = str(row["order_id"])
        month = _to_month(row["date"])
        region = str(row["region"])
        salesperson = str(row["salesperson"])
        category = str(row["category"])
        customer_type = str(row["customer_type"])
        quantity = _to_int(row["quantity"])
        amount = _to_float(row["total_amount"])

        order_ids.add(order_id)
        total_revenue += amount
        total_quantity += quantity
        revenue_by_region[region] += amount
        quantity_by_region[region] += quantity
        revenue_by_category[category] += amount
        quantity_by_category[category] += quantity
        revenue_by_salesperson[salesperson] += amount
        quantity_by_salesperson[salesperson] += quantity
        revenue_by_month[month] += amount
        quantity_by_month[month] += quantity
        revenue_by_customer[customer_type] += amount

    order_count = len(order_ids)
    average_order_value = round(total_revenue / order_count, 2) if order_count else 0.0
    region_ranking = _top_items(revenue_by_region, quantity_by_region)
    category_ranking = _top_items(revenue_by_category, quantity_by_category)
    salesperson_ranking = _top_items(revenue_by_salesperson, quantity_by_salesperson)
    monthly_trend = _trend_items(revenue_by_month, quantity_by_month)
    customer_mix = _customer_mix(revenue_by_customer, total_revenue)
    insights = _build_insights(
        total_revenue,
        order_count,
        region_ranking,
        category_ranking,
        customer_mix,
        salesperson_ranking,
        monthly_trend,
    )

    return SalesAnalysis(
        total_revenue=round(total_revenue, 2),
        order_count=order_count,
        total_quantity=total_quantity,
        average_order_value=average_order_value,
        top_regions=region_ranking,
        top_categories=category_ranking,
        monthly_trend=monthly_trend,
        region_ranking=region_ranking,
        category_ranking=category_ranking,
        salesperson_ranking=salesperson_ranking,
        customer_mix=customer_mix,
        insights=insights,
    )


def _load_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return [dict(row) for row in csv.DictReader(file)]


def _load_xlsx(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value) if value is not None else "" for value in rows[0]]
    result: list[dict[str, Any]] = []
    for values in rows[1:]:
        result.append({headers[index]: value for index, value in enumerate(values)})
    return result


def _top_items(revenue_map: dict[str, float], quantity_map: dict[str, int]) -> list[SalesMetricItem]:
    sorted_items = sorted(revenue_map.items(), key=lambda item: item[1], reverse=True)[:5]
    return [
        SalesMetricItem(name=name, revenue=round(float(revenue), 2), quantity=int(quantity_map[name]))
        for name, revenue in sorted_items
    ]


def _trend_items(revenue_map: dict[str, float], quantity_map: dict[str, int]) -> list[SalesMetricItem]:
    return [
        SalesMetricItem(name=name, revenue=round(float(revenue_map[name]), 2), quantity=int(quantity_map[name]))
        for name in sorted(revenue_map)
    ]


def _customer_mix(revenue_map: dict[str, float], total_revenue: float) -> dict[str, float]:
    if total_revenue <= 0:
        return {}
    sorted_items = sorted(revenue_map.items(), key=lambda item: item[1], reverse=True)
    return {name: round(float(value) / total_revenue * 100, 2) for name, value in sorted_items}


def _build_insights(
    total_revenue: float,
    order_count: int,
    top_regions: list[SalesMetricItem],
    top_categories: list[SalesMetricItem],
    customer_mix: dict[str, float],
    salesperson_ranking: list[SalesMetricItem],
    monthly_trend: list[SalesMetricItem],
) -> list[str]:
    insights: list[str] = []
    if top_regions:
        insights.append(f"{top_regions[0].name} 是销售额最高区域，贡献 {top_regions[0].revenue:.2f} 元。")
    if top_categories:
        insights.append(f"{top_categories[0].name} 是销售额最高品类，销售额 {top_categories[0].revenue:.2f} 元。")
    if salesperson_ranking:
        insights.append(
            f"{salesperson_ranking[0].name} 是销售额最高销售员，贡献 {salesperson_ranking[0].revenue:.2f} 元。"
        )
    if monthly_trend:
        insights.append(
            f"{monthly_trend[0].name} 月销售额 {monthly_trend[0].revenue:.2f} 元，销售件数 {monthly_trend[0].quantity}。"
        )
    if customer_mix:
        top_customer = max(customer_mix, key=customer_mix.get)
        insights.append(f"{top_customer} 贡献占比最高，占总销售额 {customer_mix[top_customer]:.2f}%。")
    if order_count > 0:
        insights.append(f"本批数据共 {order_count} 笔订单，总销售额 {total_revenue:.2f} 元。")
    return insights


def _to_float(value: Any) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _to_int(value: Any) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _to_month(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, date):
        return value.strftime("%Y-%m")
    text = str(value or "").strip()
    if len(text) >= 7:
        return text[:7]
    return "unknown"
